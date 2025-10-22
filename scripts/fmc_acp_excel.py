'''
This script will generate an excel xlsx file dump of a Cisco FMC access control policy
The output is meant to be a reference of the policy that appears similar in layout to the web GUI
The output is NOT suitable or intended to be a backup of the policy
Additionally, Users, Source Dynamic Attribute and Destination Dynamic Attribute are not accounted for but
updates to include those items should be trivial if you use those technologies
'''

import csv
import os
from typing import Any
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font
from common.fmc_lib import fmc_login, get_fmc_policies, get_fmc_rules
from fmc_config import BASE_URL

# Disable SSL warnings
import urllib3
urllib3.disable_warnings()

def get_protocol_table(protocol_file: str) -> dict[str, str]:
    """Populate protocol lookup table from CSV file

    Args:
        protocol_file (str): Path to the protocol CSV file
    Returns:
        dict[str, str]: Protocol lookup table
    """
    protocol_table: dict[str, str] = {}
    with open(protocol_file, 'r', newline="") as file:
        for num, name in csv.reader(file):
            protocol_table[str(num).strip()] = str(name).strip()
    return protocol_table

def interpret_ports(port_obj: dict, protocol_table: dict[str, str]) -> str:
    """Build a human-readable protocol/port string from an FMC rule 'port_obj'

    Args:
        port_obj (dict): port object dictionary from FMC rules
        protocol_table (dict[str, str]): protocol lookup table

    Returns:
        str: formatted port information string: TCP/443, UDP/53, ICMP/type 8, etc.
    """
    protocol_number = str(port_obj.get("protocol"))
    protocol_name = protocol_table.get(protocol_number, protocol_number).upper()
    
    if protocol_name == 'ICMP':
        icmp_type = port_obj.get("icmpType")
        return f'ICMP/type {icmp_type}' if icmp_type else 'ICMP'
    elif protocol_name in ('TCP','UDP'):
        port = port_obj.get("port")
        return f"{protocol_name}/{port}" if port else protocol_name
    return protocol_name

def output_to_excel(filename: str, columns: list[str], policyRules: list[dict[str, Any]]) -> None:
    """create excel file in similar format to FMC access control policy GUI 

    Args:
        filename (str): Path to the output Excel file
        columns (list[str]): List of column headers
        policyRules (list[dict[str, Any]]): List of policy rules to include in the Excel file
    """

    wb = Workbook() 
    ws = wb.active 
    
    #create header row in bold
    ws.append(columns)
    for i in range(len(columns)):
        ws.cell(row=1, column=i+1).font = Font(bold=True)

    #define colors for section and category separators
    darkGray = PatternFill(start_color='00BFBFBF',
                   end_color='00BFBFBF',
                   fill_type='solid')
    lightGray = PatternFill(start_color='00D9D9D9',
                   end_color='00D9D9D9',
                   fill_type='solid')
    
    #write each rule to a worksheet row
    #if a new section or category is encountered, create a separator row
    current_section = ''
    current_category = ''
    row = 2
    for rule in policyRules:
        if rule['section'] != current_section and rule['section'] != '--Undefined--':
            current_section = rule['section']
            separator = 'Section ' + current_section
            ws.cell(row=row, column=1).fill = darkGray
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(columns))
            ws.cell(row=row, column=1).value = separator
            row = row + 1
        if rule['category'] != current_category and rule['category'] != '--Undefined--':
            current_category = rule['category']
            separator = 'Category ' + current_category
            ws.cell(row=row, column=1).fill = lightGray
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(columns))
            ws.cell(row=row, column=1).value = separator
            row = row + 1
        for i in range(len(columns)):
            ws.cell(row=row, column=i+1).alignment = Alignment(wrap_text=True)
        ws.cell(row=row, column=1, value=rule['name'])
        ws.cell(row=row, column=2, value=rule['enabled'])
        ws.cell(row=row, column=3, value=rule['sourceZones'])
        ws.cell(row=row, column=4, value=rule['destinationZones'])
        ws.cell(row=row, column=5, value=rule['sourceNetworks'])
        ws.cell(row=row, column=6, value=rule['destinationNetworks'])
        ws.cell(row=row, column=7, value=rule['vlanTags'])
        ws.cell(row=row, column=8, value=rule['applications'])
        ws.cell(row=row, column=9, value=rule['sourcePorts'])
        ws.cell(row=row, column=10, value=rule['destinationPorts'])
        ws.cell(row=row, column=11, value=rule['urls'])
        ws.cell(row=row, column=12, value=rule['action'])
        ws.cell(row=row, column=13, value=rule['comments'])
        row = row + 1 

    # format column width to length of longest value plus padding
    # if the value is a multi-line string (contains \n) then format column
    # to the length of the longest piece in the multi-line string
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if '\n' in str(cell.value):
                pieces = str(cell.value).split('\n')
                for piece in pieces:
                   if len(piece) > max_length:
                        max_length = len(piece) 
            elif len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width
 
    wb.save(filename)

def join_or(items:list[str], default:str ="Any") -> str:
    """Join a list of items with a newline or return a default value.
    this formats the data for a cell in the rule
    example: ["192.168.100.1", "192.168.100.2"] = "192.168.100.1\n192.168.100.2"
    if the list is empty, that means the value in the rule is 'Any'

    Args:
        items (list): list of items to join
        default (str, optional): default value to return if items is empty. Defaults to "Any".

    Returns:
        str: joined string or default value
    """    
    return "\n".join(items) if items else default

def names(block:dict, sub:str="objects", field:str="name") -> list[str]:
    """
    Extract values from a nested list of dicts.

    Looks up `block[sub]` (expected to be a list of dict-like items) and returns
    a list containing each item's `field` value. If `sub` is missing, an empty
    list is returned. If an item lacks `field`, an empty string "" is used for
    that position.

    Args:
        block (dict): Container mapping that may hold a list under `sub`.
        sub (str, optional): Key under which the list of dicts is stored.
            Defaults to "objects".
        field (str, optional): Key to extract from each item. Defaults to "name".

    Returns:
        list[str]: Extracted values (empty strings where `field` is missing).
    """
    return [o.get(field, "") for o in block.get(sub, [])]

def values(block:dict, sub:str="literals", field:str="value") -> list[str]:
    """
    Extract values from a nested list of dicts.

    Looks up `block[sub]` (expected to be a list of dict-like items) and returns
    a list containing each item's `field`. If `sub` is missing, returns [].
    If an item lacks `field`, inserts "" for that position.

    Args:
        block (dict): Container that may hold a list under `sub`.
        sub (str, optional): Key under which the list of dicts is stored.
            Defaults to "literals".
        field (str, optional): Key to extract from each item. Defaults to "value".

    Returns:
        list[str]: Extracted values ("" where `field` is missing).
    """
    return [o.get(field, "") for o in block.get(sub, [])]

def collect_networks(block:dict) -> list[str]:
    """
    Collect network identifiers from a structure containing net object refs and ip literals.

    Combines:
      - Names from `block["objects"][i]["name"]`
      - Literal values from `block["literals"][i]["value"]`

    Delegates extraction to `names()` and `values()`, which return [] if the
    sub-list is missing and "" for missing per-item fields.

    Args:
        block (dict): Mapping that may contain "objects" (list of dicts with
            "name") and/or "literals" (list of dicts with "value").

    Returns:
        list[str]: Combined list of network names and literal values.
    """
    
    return names(block, "objects", "name") + values(block, "literals", "value")

def collect_ports(block:dict, protocol_table:dict[str, str]) -> list[str]:
    """
    Collect port/service identifiers from a structure containing object refs and
    port literals.

    Combines:
      - Named services from `block["objects"][i]["name"]` (via `names()`).
      - Parsed literals from `block["literals"][i]`, each rendered by
        `interpret_ports(literal, protocol_table)` into a human-readable token
        (e.g., "tcp/443", "udp/53").

    Missing sublists are treated as empty. Per-item missing fields in the
    "objects" list become "" (per `names()` behavior).

    Args:
        block (dict): May contain "objects" (list of dicts with "name") and/or
            "literals" (list of dicts describing protocol/ports/ranges).
        protocol_table (Mapping): Lookup used by `interpret_ports` to map protocol
            identifiers (e.g., 6 → "tcp", 17 → "udp") to names.

    Returns:
        list[str]: Combined list of service names and interpreted port literals.
    """

    out = names(block, "objects", "name")
    out += [interpret_ports(lit, protocol_table) for lit in block.get("literals", [])]
    return out

def collect_vlans(block:dict) -> list[str]:
    """
    Collect VLAN identifiers from a structure containing object refs and literal ranges.

    Combines:
      - Names from `block["objects"][i]["name"]` (via `names()`).
      - Ranges from `block["literals"][i]` formatted as `"startTag-endTag"`.

    Missing sublists are treated as empty. If a literal lacks `startTag` or
    `endTag`, the resulting string will include "None" for the missing bound.

    Args:
        block (dict): Mapping that may contain "objects" (list of dicts with
            "name") and/or "literals" (list of dicts with "startTag"/"endTag").

    Returns:
        list[str]: Combined list of VLAN names and VLAN range strings.
    """
    
    out = names(block, "objects", "name")
    out += [f"{lit.get('startTag')}-{lit.get('endTag')}" for lit in block.get("literals", [])]
    return out

def collect_urls(block:dict) -> list[str]:
    """
    Collect URL match criteria from category/reputation pairs and literal URLs.

    Combines:
      - Category/reputation tokens from `block["urlCategoriesWithReputation"]`,
        formatted as "<category>/<reputation>". Items without a category name
        are skipped; a missing reputation yields an empty string after the slash.
      - Literal URLs from `block["literals"][i]["url"]`.

    Missing sublists are treated as empty.

    Args:
        block (dict): May contain:
            - "urlCategoriesWithReputation": list of dicts with
              {"category": {"name": str}, "reputation": Optional[str]}
            - "literals": list of dicts with {"url": str}

    Returns:
        list[str]: Combined list of "<category>/<reputation>" tokens and literal URLs.
    """
    
    out = [
        f"{it['category']['name']}/{it.get('reputation','')}"
        for it in block.get("urlCategoriesWithReputation", [])
        if it.get("category", {}).get("name")
    ]
    out += [u.get("url", "") for u in block.get("literals", [])]
    return out

def transform_rules(rules: list[dict[str, Any]], protocol_table: dict[str, str]) -> list[dict[str, Any]]:
    """
    Normalize an FMC policy rules payload into report-friendly rows.

    Iterates `rules` and builds a list of dicts with consistent keys:
    `enabled`, `name`, `action`, `section`, `category`, `comments`,
    `sourceZones`, `destinationZones`, `sourceNetworks`, `destinationNetworks`,
    `vlanTags`, `applications`, `sourcePorts`, `destinationPorts`, `urls`.

    Field construction:
      - Zones/Networks/Applications/VLANs/Ports/URLs are assembled via helper
        collectors (`names`, `collect_networks`, `collect_vlans`,
        `collect_ports`, `collect_urls`) and then newline-joined via `join_or`.
      - Empty aggregates default to "Any" (per `join_or`), except `comments`,
        which defaults to an empty string "" when absent.
      - Port literals are rendered by `interpret_ports(..., protocol_table)`; the
        mapping should translate protocol numbers (e.g., "6" → "tcp") to names.

    Filtering:
      - If `include_one_offs` is False (default), rules whose
        `metadata.category` contains "one-off" (case-insensitive) are skipped.

    Args:
        rules (list[dict[str, Any]]): An FMC-style response containing a list of rule dicts.
        protocol_table (dict[str, str]): Protocol-number → name mapping used by
            `interpret_ports` within `collect_ports`.
    Returns:
        list[dict]: Normalized rules ready for export/display, with the keys listed above.
    """
    
    out_rules = []
    for rule in rules:
        meta = rule.get("metadata", {})
        # skip any rules in one-off section of the policy
        out = {
            "enabled":  rule.get("enabled"),
            "name":     rule.get("name"),
            "action":   rule.get("action"),
            "section":  meta.get("section"),
            "category": meta.get("category"),
            "comments": "\n".join(c.get("comment","") for c in rule.get("commentHistoryList", [])),
            "sourceZones":      join_or(names(rule.get("sourceZones", {}))),
            "destinationZones": join_or(names(rule.get("destinationZones", {}))),
            "sourceNetworks":      join_or(collect_networks(rule.get("sourceNetworks", {}))),
            "destinationNetworks": join_or(collect_networks(rule.get("destinationNetworks", {}))),
            "vlanTags":        join_or(collect_vlans(rule.get("vlanTags", {}))),
            "applications":    join_or(names(rule.get("applications", {}), "applications", "name")),
            "sourcePorts":     join_or(collect_ports(rule.get("sourcePorts", {}), protocol_table)),
            "destinationPorts":join_or(collect_ports(rule.get("destinationPorts", {}), protocol_table)),
            "urls":            join_or(collect_urls(rule.get("urls", {}))),
        }

        # Match your exact original default for comments: empty when missing
        if "commentHistoryList" not in rule:
            out["comments"] = ""

        out_rules.append(out)

    return out_rules

def main() -> None:

    # vars
    base_url = BASE_URL
    protocol_file = 'data/protocols.csv'  #based on iana data
    protocol_table = get_protocol_table(protocol_file)
    output_path = f'output'  # location to same excel file to
    #columns to be output to file in this order, each column matching a dictionary key for an item in policy_rules
    columns = ['name', 'enabled', 'sourceZones', 'destinationZones', 'sourceNetworks', 'destinationNetworks', 'vlanTags', 
                   'applications', 'sourcePorts', 'destinationPorts', 'urls', 'action', 'comments']
    
    # ensure the output path exists
    # no try, just blow up since there's no point in continuing
    print("Ensuring output directory exists...")
    os.makedirs(output_path, exist_ok=True)
    print(f"Directory '{output_path}' created successfully or already exists.")

    # login and retrieve token and DUUID
    result = fmc_login(base_url)
    token = result.get('X-auth-access-token')
    DUUID = result.get('DOMAIN_UUID')
    
    # get the list of access control policies in FMC
    print("Retrieving access control policies...")
    policies = get_fmc_policies(base_url, token, DUUID)
    if not policies:
        raise Exception("No policies found or error retrieving policies")
    
    # prompt for input on which policy to examine
    print('\nPolicies found')
    for i, item in enumerate(policies, 1):
        print(f"[ {i} ] {item['name']}")
    entry = input('Enter the number of the policy you want to export: ')
    try:
        entry = int(entry)
    except ValueError:
        raise ValueError("Invalid entry: must be a number.")
    if entry < 1 or entry > len(policies):
        raise ValueError("Invalid entry: out of range.")
    policy = policies[entry -1]
    print()

    # get the list of rules associated with the policy
    print("Retrieving rules for access control policy:", policy["name"])
    rules = get_fmc_rules(base_url, token, DUUID, policy['id'])
    if not rules:
        raise Exception(f'No rules were retrieved for access control policy {policy["name"]}')

    # transform the rules into a format suitable for Excel output
    print("Transforming rules for Excel output...")
    policy_rules = transform_rules(rules, protocol_table)
    
    # output to excel
    timestamp = datetime.now().strftime('%Y%m%d-%H%M%S')
    safe_name = "".join(c for c in policy["name"] if c.isalnum() or c in (" ", "_", "-")).rstrip()
    excel_filename = f'FMC-ACP-{safe_name}-{timestamp}.xlsx'
    excel_file = os.path.join(output_path, excel_filename)
    print("Generating Excel file...")
    output_to_excel(excel_file, columns, policy_rules)
    print(f'XLSX output for access control policy {policy["name"]} complete')

if __name__ == "__main__":
    main()

